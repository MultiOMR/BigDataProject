'''
Created on 10/11/2014

@organization: Lancaster University & University of Leeds
@version: 1.0
Created on 11/12/2014

@author: Victor Padilla
@contact: v.padilla@lancaster.ac.uk

Class for writing data in .xlsx format
It allows writing the results comparing with the ground in .xls

'''
import xlsxwriter 
from openpyxl import load_workbook

class ExcellData:
    def saveData(self,data,files,percentages):
        '''
        save the data from one part (one movement)
        in the "result.xlsx" file
        
        usage:
        
        ed=ExcellData()
        ff=FilesFunctions()
        OMRs=ff.getOMRs(path)
        for omr in OMRs:
            OMRs_compare=[]
            OMRs_compare.append(groundparsed)
            OMRs_compare.append(omr)

            percentage,errors,scoreWithErrors= pg.getSimilarity(OMRs_compare,0)
            ErrorsMatrix.append(errors)
            percentages.append(percentage)

        ed.saveData(ErrorsMatrix,files,percentages) 
        '''
         
        path=self.__getPathStepsBack(files[0],2)
        wb=xlsxwriter.Workbook(path+'\\Result\\result.xlsx')
        ws=wb.add_worksheet()
        c=0
        ws.write(0, 0,str(files[0]))
        
        for col in data:
            r=2
            ws.write(1, c,str(percentages[c]))
            fArray=files[c].split('\\')
            ws.write(r, c,str(fArray[-1]))
            for row in col:
                r+=1
                ws.write(r, c,str(row))
            c+=1   
        wb.close()
        
    def saveGlobalData(self,data,path,betterOMRIds,files):
        '''
        Takes the individual parts and store the result 
        in "resultGeneral.xlsx"
        
        usage:
        ed=ExcellData()
        ed.saveGlobalData(percentagesArray,dirGeneral,betterOMRIds,files) 
        
        files=['file1.xml','file2.xml','file3.xml','file4.xml']
        
        data=[[0.7 ,0.8, 0.5, 0.6] 
              [0.7 ,0.8, 0.5, 0.6]
              [0.7 ,0.8, 0.5, 0.6]]
              
        betterOMRIds=[[0,1,2]
                      [2,3,1]
                      [1,0,2]]
        
        ed=ExcellData()
        ed.saveGlobalData(percentagesArray,dirGeneral,betterOMRIds,files) 
        '''
        
        
        wb=xlsxwriter.Workbook(path+'\\parts\\resultGeneral.xlsx')
        ws=wb.add_worksheet()
        r=0
        ws.write(0, 1,str(files[0]))
        print betterOMRIds
        for row in data:
            c=0
            myBetterOMR=betterOMRIds[r]
            for col in row:
                format1 = wb.add_format()
                format1.set_font_color('black')
                if str(c-1)+"\n" in myBetterOMR:
                    format1.set_font_color('red')
                fArray=files[c].split('\\')
                ws.write(0, c,str(fArray[-1]))
                ws.write(r+1, c,float(col),format1)
                c+=1
            r+=1   
        wb.close()
    
    def processResultGeneral(self,f): 
        '''
        This function takes one xls file and get the percentage of accuracy of each omr
        Returns in this order  S2Out,CPOut,PSOut,SEOut,SSOut
        '''
        wb = load_workbook(f)
        ws = wb['Sheet1']
        rows=ws.iter_rows()
        S2=0
        iS2=0
        CP=0
        iCP=0
        PS=0
        iPS=0
        SE=0
        iSE=0
        SS=0
        iSS=0
        indexRow=0
        headRow=[]
        for row in rows:
            if (indexRow==0):
                headRow=row
            else:
                for cell in row:
                    col=row.index(cell)
                    strHead=headRow[col].value.upper()
                    if '.S2.XML' in strHead:
                        try:
                            S2+=cell.value
                            iS2+=1
                        except:
                            print "error"
                            
                    if '.CP.XML' in strHead:
                        try:
                            CP+=cell.value
                            iCP+=1
                        except:
                            print "error"
                    if '.PS.XML' in strHead:
                        try:
                            PS+=cell.value
                            iPS+=1
                        except:
                            print "error"
                    if '.SE.XML' in strHead:
                        try:
                            SE+=cell.value
                            iSE+=1
                        except:
                            print "error"
                    if '.SS.XML' in strHead:
                        try:
                            SS+=cell.value
                            iSS+=1
                        except:
                            print "error"
            indexRow+=1
        S2Out=0
        CPOut=0
        PSOut=0
        SEOut=0
        SSOut=0
        if iS2>0:
            S2Out= S2/iS2
        if iCP>0:
            CPOut= CP/iCP
        if iPS>0:
            PSOut=PS/iPS
        if SE>0:
            SEOut=SE/iSE
        if SS>0:
            SSOut=SS/iSS
        return S2Out,CPOut,PSOut,SEOut,SSOut
    
    def writeFinalXLS(self, rootDir,S2,CP,PS,SE,SS):
        '''
        Writes the final.xlsx upon the inputs S2,CP,PS,SE,SS
        '''
        wb=xlsxwriter.Workbook(rootDir+'\\final.xlsx')
        ws=wb.add_worksheet()
        ws.write(0, 0,"S2")
        ws.write(0, 1,"CP")
        ws.write(0, 2,"PS")
        ws.write(0, 3,"SE")
        ws.write(0, 4,"SS")
        ws.write(1, 0,S2)
        ws.write(1, 1,CP)
        ws.write(1, 2,PS)
        ws.write(1, 3,SE)
        ws.write(1, 4,SS)
        
        wb.close()
        
    def __getPathStepsBack(self,f,stepsBack):
        '''
        Private function: takes n steps back in the path from a file
        '''
        fArray=f.split('\\')
        path=''
        for i in range(len(fArray)-stepsBack):
            path+=fArray[i]+"/"
        return path
    